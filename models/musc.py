import os
import sys
import numpy as np
import torch
import torch.nn.functional as F
sys.path.append('./models/backbone')

import datasets.mvtec as mvtec
from datasets.mvtec import _CLASSNAMES as _CLASSNAMES_mvtec_ad
import datasets.visa as visa
from datasets.visa import _CLASSNAMES as _CLASSNAMES_visa
import datasets.btad as btad
from datasets.btad import _CLASSNAMES as _CLASSNAMES_btad

import models.backbone.open_clip as open_clip
import models.backbone._backbones as _backbones
from models.modules._LNAMD import LNAMD
from models.modules._MSM import MSM
from models.modules._RsCIN import RsCIN
from utils.metrics import compute_metrics
from openpyxl import Workbook
from tqdm import tqdm
import pickle
import time
import cv2

import warnings
warnings.filterwarnings("ignore")


class MuSc():
    def __init__(self, cfg, seed=0):
        self.cfg = cfg
        self.seed = seed
        self.device = torch.device("cuda:{}".format(cfg['device']) if torch.cuda.is_available() else "cpu")

        self.path = cfg['datasets']['data_path']
        self.dataset = cfg['datasets']['dataset_name']
        self.vis = cfg['testing']['vis']
        self.vis_type = cfg['testing']['vis_type']
        self.save_excel = cfg['testing']['save_excel']
        # the categories to be tested
        self.categories = cfg['datasets']['class_name']
        if isinstance(self.categories, str):
            if self.categories.lower() == 'all':
                if self.dataset == 'visa':
                    self.categories = _CLASSNAMES_visa
                elif self.dataset == 'mvtec_ad':
                    self.categories = _CLASSNAMES_mvtec_ad
                elif self.dataset == 'btad':
                    self.categories = _CLASSNAMES_btad
            else:
                self.categories = [self.categories]

        self.model_name = cfg['models']['backbone_name']
        self.image_size = cfg['datasets']['img_resize']
        self.batch_size = cfg['models']['batch_size']
        self.pretrained = cfg['models']['pretrained']
        self.features_list = [l+1 for l in cfg['models']['feature_layers']]
        self.divide_num = cfg['datasets']['divide_num']
        self.r_list = cfg['models']['r_list']
        self.output_dir = os.path.join(cfg['testing']['output_dir'], self.dataset, self.model_name, 'imagesize{}'.format(self.image_size))
        os.makedirs(self.output_dir, exist_ok=True)
        self.load_backbone()


    def load_backbone(self):
        if 'dino' in self.model_name:
            # dino or dino_v2
            self.dino_model = _backbones.load(self.model_name)
            self.dino_model.to(self.device)
            self.preprocess = None
        else:
            # clip
            self.clip_model, _, self.preprocess = open_clip.create_model_and_transforms(self.model_name, self.image_size, pretrained=self.pretrained)
            self.clip_model.to(self.device)


    def load_datasets(self, category, divide_num=1, divide_iter=0):
        # dataloader
        if self.dataset == 'visa':
            test_dataset = visa.VisaDataset(source=self.path, split=visa.DatasetSplit.TEST,
                                            classname=category, resize=self.image_size, imagesize=self.image_size, clip_transformer=self.preprocess,
                                                divide_num=divide_num, divide_iter=divide_iter, random_seed=self.seed)
        elif self.dataset == 'mvtec_ad':
            test_dataset = mvtec.MVTecDataset(source=self.path, split=mvtec.DatasetSplit.TEST,
                                            classname=category, resize=self.image_size, imagesize=self.image_size, clip_transformer=self.preprocess,
                                                divide_num=divide_num, divide_iter=divide_iter, random_seed=self.seed)
        elif self.dataset == 'btad':
            test_dataset = btad.BTADDataset(source=self.path, split=btad.DatasetSplit.TEST,
                                            classname=category, resize=self.image_size, imagesize=self.image_size, clip_transformer=self.preprocess,
                                                divide_num=divide_num, divide_iter=divide_iter, random_seed=self.seed)
        return test_dataset


    def visualization(self, image_path_list, gt_list, pr_px, category):
        def normalization01(img):
            return (img - img.min()) / (img.max() - img.min())
        if self.vis_type == 'single_norm':
            # normalized per image
            for i, path in enumerate(image_path_list):
                anomaly_type = path.split('/')[-2]
                img_name = path.split('/')[-1]
                if anomaly_type not in ['good', 'Normal', 'ok'] and gt_list[i] != 0:
                    save_path = os.path.join(self.output_dir, category, anomaly_type)
                    os.makedirs(save_path, exist_ok=True)
                    save_path = os.path.join(save_path, img_name)
                    anomaly_map = pr_px[i].squeeze()
                    anomaly_map = normalization01(anomaly_map)*255
                    anomaly_map = cv2.applyColorMap(anomaly_map.astype(np.uint8), cv2.COLORMAP_JET)
                    cv2.imwrite(save_path, anomaly_map)
        else:
            # normalized all image
            pr_px = normalization01(pr_px)
            for i, path in enumerate(image_path_list):
                anomaly_type = path.split('/')[-2]
                img_name = path.split('/')[-1]
                save_path = os.path.join(self.output_dir, category, anomaly_type)
                os.makedirs(save_path, exist_ok=True)
                save_path = os.path.join(save_path, img_name)
                anomaly_map = pr_px[i].squeeze()
                anomaly_map *= 255
                anomaly_map = cv2.applyColorMap(anomaly_map.astype(np.uint8), cv2.COLORMAP_JET)
                cv2.imwrite(save_path, anomaly_map)


    def make_category_data(self, category):
        print(category)

        # divide sub-datasets
        divide_num = self.divide_num
        anomaly_maps = torch.tensor([]).double()
        gt_list = []
        img_masks = []
        class_tokens = []
        image_path_list = []
        start_time_all = time.time()
        dataset_num = 0
        for divide_iter in range(divide_num):
            test_dataset = self.load_datasets(category, divide_num=divide_num, divide_iter=divide_iter)
            test_dataloader = torch.utils.data.DataLoader(
                test_dataset,
                batch_size=self.batch_size,
                shuffle=False,
                num_workers=0,
                pin_memory=True,
            )
            
            # extract features
            patch_tokens_list = []
            subset_num = len(test_dataset)
            dataset_num += subset_num
            start_time = time.time()
            for image_info in tqdm(test_dataloader):
            # for image_info in test_dataloader:
                if isinstance(image_info, dict):
                    image = image_info["image"]
                    image_path_list.extend(image_info["image_path"])
                    img_masks.append(image_info["mask"])
                    gt_list.extend(list(image_info["is_anomaly"].numpy()))
                with torch.no_grad(), torch.cuda.amp.autocast():
                    input_image = image.to(torch.float).to(self.device)
                    if 'dinov2' in self.model_name:
                        patch_tokens = self.dino_model.get_intermediate_layers(x=input_image, n=[l-1 for l in self.features_list], return_class_token=False)
                        image_features = self.dino_model(input_image)
                        patch_tokens = [patch_tokens[l].cpu() for l in range(len(self.features_list))]
                        fake_cls = [torch.zeros_like(p)[:, 0:1, :] for p in patch_tokens]
                        patch_tokens = [torch.cat([fake_cls[i], patch_tokens[i]], dim=1) for i in range(len(patch_tokens))]
                    elif 'dino' in self.model_name:
                        patch_tokens_all = self.dino_model.get_intermediate_layers(x=input_image, n=max(self.features_list))
                        image_features = self.dino_model(input_image)
                        patch_tokens = [patch_tokens_all[l-1].cpu() for l in self.features_list]
                    else: # clip
                        image_features, patch_tokens = self.clip_model.encode_image(input_image, self.features_list)
                        image_features /= image_features.norm(dim=-1, keepdim=True)
                        patch_tokens = [patch_tokens[l].cpu() for l in range(len(self.features_list))]
                image_features = [image_features[bi].squeeze().cpu().numpy() for bi in range(image_features.shape[0])]
                class_tokens.extend(image_features)
                patch_tokens_list.append(patch_tokens)  # (B, L+1, C)
            end_time = time.time()
            print('extract time: {}ms per image'.format((end_time-start_time)*1000/subset_num))
            
            # LNAMD
            feature_dim = patch_tokens_list[0][0].shape[-1]
            anomaly_maps_r = torch.tensor([]).double()
            for r in self.r_list:
                start_time = time.time()
                print('aggregation degree: {}'.format(r))
                LNAMD_r = LNAMD(device=self.device, r=r, feature_dim=feature_dim, feature_layer=self.features_list)
                Z_layers = {}
                for im in range(len(patch_tokens_list)):
                    patch_tokens = [p.to(self.device) for p in patch_tokens_list[im]]
                    with torch.no_grad(), torch.cuda.amp.autocast():
                        features = LNAMD_r._embed(patch_tokens)
                        features /= features.norm(dim=-1, keepdim=True)
                        for l in range(len(self.features_list)):
                            # save the aggregated features
                            if str(l) not in Z_layers.keys():
                                Z_layers[str(l)] = []
                            Z_layers[str(l)].append(features[:, :, l, :])
                end_time = time.time()
                print('LNAMD-{}: {}ms per image'.format(r, (end_time-start_time)*1000/subset_num))

                # MSM
                anomaly_maps_l = torch.tensor([]).double()
                start_time = time.time()
                for l in Z_layers.keys():
                    # different layers
                    Z = torch.cat(Z_layers[l], dim=0).to(self.device) # (N, L, C)
                    print('layer-{} mutual scoring...'.format(l))
                    anomaly_maps_msm = MSM(Z=Z, device=self.device, topmin_min=0, topmin_max=0.3)
                    anomaly_maps_l = torch.cat((anomaly_maps_l, anomaly_maps_msm.unsqueeze(0).cpu()), dim=0)
                    torch.cuda.empty_cache()
                anomaly_maps_l = torch.mean(anomaly_maps_l, 0)
                anomaly_maps_r = torch.cat((anomaly_maps_r, anomaly_maps_l.unsqueeze(0)), dim=0)
                end_time = time.time()
                print('MSM: {}ms per image'.format((end_time-start_time)*1000/subset_num))
            anomaly_maps_iter = torch.mean(anomaly_maps_r, 0).to(self.device)
            del anomaly_maps_r
            torch.cuda.empty_cache()

            # interpolate
            B, L = anomaly_maps_iter.shape
            H = int(np.sqrt(L))
            anomaly_maps_iter = F.interpolate(anomaly_maps_iter.view(B, 1, H, H),
                                        size=self.image_size, mode='bilinear', align_corners=True)
            anomaly_maps = torch.cat((anomaly_maps, anomaly_maps_iter.cpu()), dim=0)

        # save image features for optimizing classification
        # cls_save_path = os.path.join('./image_features/{}_{}.dat'.format(dataset, category))
        # with open(cls_save_path, 'wb') as f:
        #     pickle.dump([np.array(class_tokens)], f)
        end_time_all = time.time()
        print('MuSc: {}ms per image'.format((end_time_all-start_time_all)*1000/dataset_num))

        anomaly_maps = anomaly_maps.cpu().numpy()
        torch.cuda.empty_cache()

        B = anomaly_maps.shape[0]   # the number of unlabeled test images
        ac_score = np.array(anomaly_maps).reshape(B, -1).max(-1)
        # RsCIN
        if self.dataset == 'visa':
            k_score = [1, 8, 9]
        elif self.dataset == 'mvtec_ad':
            k_score = [1, 2, 3]
        else:
            k_score = [1, 2, 3]
        scores_cls = RsCIN(ac_score, class_tokens, k_list=k_score)

        print('computing metrics...')
        pr_sp = np.array(scores_cls)
        gt_sp = np.array(gt_list)
        gt_px = torch.cat(img_masks, dim=0).numpy().astype(np.int32)
        pr_px = np.array(anomaly_maps)
        image_metric, pixel_metric = compute_metrics(gt_sp, pr_sp, gt_px, pr_px)
        auroc_sp, f1_sp, ap_sp = image_metric
        auroc_px, f1_px, ap_px, aupro = pixel_metric
        print(category)
        print('image-level, auroc:{}, f1:{}, ap:{}'.format(auroc_sp*100, f1_sp*100, ap_sp*100))
        print('pixel-level, auroc:{}, f1:{}, ap:{}, aupro:{}'.format(auroc_px*100, f1_px*100, ap_px*100, aupro*100))

        if self.vis:
            print('visualization...')
            self.visualization(image_path_list, gt_list, pr_px, category)
    
        return image_metric, pixel_metric


    def main(self):
        auroc_sp_ls = []
        f1_sp_ls = []
        ap_sp_ls = []
        auroc_px_ls = []
        f1_px_ls = []
        ap_px_ls = []
        aupro_ls = []
        for category in self.categories:
            image_metric, pixel_metric = self.make_category_data(category=category,)
            auroc_sp, f1_sp, ap_sp = image_metric
            auroc_px, f1_px, ap_px, aupro = pixel_metric
            auroc_sp_ls.append(auroc_sp)
            f1_sp_ls.append(f1_sp)
            ap_sp_ls.append(ap_sp)
            auroc_px_ls.append(auroc_px)
            f1_px_ls.append(f1_px)
            ap_px_ls.append(ap_px)
            aupro_ls.append(aupro)
        # mean
        auroc_sp_mean = sum(auroc_sp_ls) / len(auroc_sp_ls)
        f1_sp_mean = sum(f1_sp_ls) / len(f1_sp_ls)
        ap_sp_mean = sum(ap_sp_ls) / len(ap_sp_ls)
        auroc_px_mean = sum(auroc_px_ls) / len(auroc_px_ls)
        f1_px_mean = sum(f1_px_ls) / len(f1_px_ls)
        ap_px_mean = sum(ap_px_ls) / len(ap_px_ls)
        aupro_mean = sum(aupro_ls) / len(aupro_ls)

        for i, category in enumerate(self.categories):
            print(category)
            print('image-level, auroc:{}, f1:{}, ap:{}'.format(auroc_sp_ls[i]*100, f1_sp_ls[i]*100, ap_sp_ls[i]*100))
            print('pixel-level, auroc:{}, f1:{}, ap:{}, aupro:{}'.format(auroc_px_ls[i]*100, f1_px_ls[i]*100, ap_px_ls[i]*100, aupro_ls[i]*100))
        print('mean')
        print('image-level, auroc:{}, f1:{}, ap:{}'.format(auroc_sp_mean*100, f1_sp_mean*100, ap_sp_mean*100))
        print('pixel-level, auroc:{}, f1:{}, ap:{}, aupro:{}'.format(auroc_px_mean*100, f1_px_mean*100, ap_px_mean*100, aupro_mean*100))
        
        # save in excel
        if self.save_excel:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "MuSc_results"
            sheet.cell(row=1,column=2,value='auroc_px')
            sheet.cell(row=1,column=3,value='f1_px')
            sheet.cell(row=1,column=4,value='ap_px')
            sheet.cell(row=1,column=5,value='aupro')
            sheet.cell(row=1,column=6,value='auroc_sp')
            sheet.cell(row=1,column=7,value='f1_sp')
            sheet.cell(row=1,column=8,value='ap_sp')
            for col_index in range(2):
                for row_index in range(len(self.categories)):
                    if col_index == 0:
                        sheet.cell(row=row_index+2,column=col_index+1,value=self.categories[row_index])
                    else:
                        sheet.cell(row=row_index+2,column=col_index+1,value=auroc_px_ls[row_index]*100)
                        sheet.cell(row=row_index+2,column=col_index+2,value=f1_px_ls[row_index]*100)
                        sheet.cell(row=row_index+2,column=col_index+3,value=ap_px_ls[row_index]*100)
                        sheet.cell(row=row_index+2,column=col_index+4,value=aupro_ls[row_index]*100)
                        sheet.cell(row=row_index+2,column=col_index+5,value=auroc_sp_ls[row_index]*100)
                        sheet.cell(row=row_index+2,column=col_index+6,value=f1_sp_ls[row_index]*100)
                        sheet.cell(row=row_index+2,column=col_index+7,value=ap_sp_ls[row_index]*100)
                    if row_index == len(self.categories)-1:
                        if col_index == 0:
                            sheet.cell(row=row_index+3,column=col_index+1,value='mean')
                        else:
                            sheet.cell(row=row_index+3,column=col_index+1,value=auroc_px_mean*100)
                            sheet.cell(row=row_index+3,column=col_index+2,value=f1_px_mean*100)
                            sheet.cell(row=row_index+3,column=col_index+3,value=ap_px_mean*100)
                            sheet.cell(row=row_index+3,column=col_index+4,value=aupro_mean*100)
                            sheet.cell(row=row_index+3,column=col_index+5,value=auroc_sp_mean*100)
                            sheet.cell(row=row_index+3,column=col_index+6,value=f1_sp_mean*100)
                            sheet.cell(row=row_index+3,column=col_index+7,value=ap_sp_mean*100)
            workbook.save(os.path.join(self.output_dir, 'results.xlsx'))


